import openpyxl

wb = openpyxl.load_workbook("attendance_sheet.xlsx", data_only=True)  # Enable formula value reading
subjects = wb.sheetnames

while True:
    print()
    open_sheet = input(f"Enter the name of the subject {subjects}: ").lower()
    if open_sheet in subjects:
        sheet_name = wb[open_sheet]  # Actual worksheet object
        break
    else:
        print(f"❌ Invalid subject name.")
        if input("Do you want to continue? (y/n): ").lower() != "y":
            exit()

low_attendance_list = []

row = 5
while True:
    percentage = sheet_name.cell(row=row, column=27).value  # Column AA = 27
    name_or_roll = sheet_name.cell(row=row, column=2).value  # Column 2 = Name or Roll Number

    if name_or_roll is None:
        break  # Stop at first empty row

    if percentage is not None and isinstance(percentage, (int, float)) and percentage < 75:
        low_attendance_list.append(name_or_roll)

    row += 1

print()
print("students with low attendance >75%")
for i in low_attendance_list:
    print()
    print(i)
