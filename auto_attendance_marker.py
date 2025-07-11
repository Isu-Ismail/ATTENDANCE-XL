import os
import time
import openpyxl as xl
from datetime import date
from openpyxl.styles import Font

today = date.today()
open_file = "attendance_sheet.xlsx"
wb = xl.load_workbook(open_file)

# Wait until the Excel file is closed
def wait_until_file_closed(filename=open_file):
    while True:
        try:
            with open(filename, "a"):
                return
        except PermissionError:
            print(f"⏳ '{filename}' is currently open. Please close it to proceed.")
            time.sleep(3)

# Find next empty column
def selectedCol(sheet):
    col = 3
    while sheet.cell(row=3, column=col).value:
        col += 1
    return col

#  Count total student rows
def selectedRow(sheet, start_row=5):
    row = start_row
    while sheet.cell(row=row, column=1).value:
        row += 1
    return row - 4

#  Convert range input like 1-5,8 to list of integers
def parse_roll_numbers(input_string):
    result = set()
    parts = input_string.split(',')
    for part in parts:
        part = part.strip()
        if '-' in part:
            start, end = part.split('-')
            result.update(range(int(start), int(end) + 1))
        else:
            result.add(int(part))
    return sorted(result)

# Mark all as present and add summary
def full_present(sheet, selected_column, number_of_hours, filename=open_file):
    sheet.cell(3, selected_column).value = number_of_hours
    sheet.cell(2, selected_column).value = today.strftime("%d-%m-%Y")

    total_rows = selectedRow(sheet)
    for i in range(5, total_rows + 4):
        sheet.cell(i, selected_column).value = "P"

    # Add summary
    summary_row = total_rows + 6
    font_style = Font(name='Times New Roman', size=10, bold=True)
    sheet.cell(summary_row, selected_column).value = f"Total Present: {total_rows}"
    sheet.cell(summary_row + 1, selected_column).value = f"Total Absent: 0"
    sheet.cell(summary_row, selected_column).font = font_style
    sheet.cell(summary_row + 1, selected_column).font = font_style

    wb.save(filename)

#  Mark attendance using absentees list and add summary
def auto_attendance_marker(sheet, selected_column, number_of_hours, absentees, filename=open_file):
    sheet.cell(3, selected_column).value = number_of_hours
    sheet.cell(2, selected_column).value = today.strftime("%d-%m-%Y")

    total_rows = selectedRow(sheet)
    present_count = 0
    absent_count = 0

    for i in range(5, total_rows + 4):
        roll = sheet.cell(i, 1).value
        cell = sheet.cell(i, selected_column)
        if roll in absentees:
            cell.value = "A"
            absent_count += 1
        else:
            cell.value = "P"
            present_count += 1

    # Add summary
    summary_row = total_rows + 6
    font_style = Font(name='Times New Roman', size=10, bold=True)
    sheet.cell(summary_row, selected_column).value = f"Total Present: {present_count}"
    sheet.cell(summary_row + 1, selected_column).value = f"Total Absent: {absent_count}"
    sheet.cell(summary_row, selected_column).font = font_style
    sheet.cell(summary_row + 1, selected_column).font = font_style

    wb.save(filename)

# --- Main Execution ---
wait_until_file_closed(open_file)
print(f"✅ '{open_file}' is closed. Proceeding...")

subjects = wb.sheetnames

#  Ask for subject name
while True:
    print()
    open_sheet = input(f"Enter the name of the subject {subjects}: ").lower()
    if open_sheet in subjects:
        sheet_name = wb[open_sheet]
        break
    else:
        print(f"❌ Invalid subject name.")
        if input("Do you want to continue? (y/n): ").lower() != "y":
            exit()

# Ask for absentees or full present
while True:
    print()
    user_input = input(
        "Enter absent roll numbers (e.g. 2,4-7,9) or 0 for full present:\n"
        "👉 Or use `!` followed by present numbers to mark all others as absent (e.g., !1-5,7)\n> "
    ).strip()

    if user_input == "0":
        while True:
            try:
                print()
                number_hours = int(input("Enter number of hours class was taken (max: 4): "))
                selected_col = selectedCol(sheet_name)
                if selected_col < 4:
                    raise ValueError("Select column >= 4")
                full_present(sheet_name, selected_col, number_hours)
                print("✅ Full present marked.")
                exit()
            except ValueError as e:
                print(f"❌ {e}")
                if input("Do you want to continue? (y/n): ").lower() != "y":
                    exit()

    else:
        try:
            total_students = selectedRow(sheet_name)

            if user_input.startswith("!"):
                present_rolls = parse_roll_numbers(user_input[1:])
                absentees_list = [r for r in range(1, total_students + 1) if r not in present_rolls]
            else:
                absentees_list = parse_roll_numbers(user_input)

        except ValueError:
            print("❌ Error: Only integers or ranges separated by commas are allowed.")
            if input("Do you want to continue? (y/n): ").lower() != "y":
                exit()
            continue

        invalid_rolls = [r for r in absentees_list if r < 1 or r > total_students]
        if invalid_rolls:
            print("❌ Invalid roll numbers:", ", ".join(map(str, invalid_rolls)))
            if input("Do you want to continue? (y/n): ").lower() != "y":
                exit()
            continue

        while True:
            try:
                print()
                number_hours = int(input("Enter number of hours class was taken (max: 4): "))
                selected_col = selectedCol(sheet_name)
                if selected_col < 4:
                    raise ValueError("Select column >= 4")
                auto_attendance_marker(sheet_name, selected_col, number_hours, absentees_list)
                print("✅ Attendance marked with absentees.")
                exit()
            except ValueError as e:
                print(f"❌ {e}")
                if input("Do you want to continue? (y/n): ").lower() != "y":
                    exit()
